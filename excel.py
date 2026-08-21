import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# --- FUNÇÕES AUXILIARES DE FORMATAÇÃO ---

def converter_para_numero(valor_str):
    """Converte '1.234,56' para 1234.56. Se não for número, retorna 0.0"""
    if not valor_str or valor_str in ("-", "Consultar", "Ver Detalhes"):
        return 0.0
    try:
        # Remove os pontos de milhar e troca vírgula por ponto decimal
        limpo = valor_str.replace('.', '').replace(',', '.')
        return float(limpo)
    except:
        return 0.0

def estilizar_cabecalho(ws, headers):
    ws.append(headers)
    fundo = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate 900
    fonte = Font(color="FFFFFF", bold=True)
    borda = Border(bottom=Side(style='thick', color="000000"))
    
    for i, cell in enumerate(ws[1], 1):
        cell.font = fonte
        cell.fill = fundo
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
    
    # Congela a primeira linha para o cabeçalho acompanhar a rolagem
    ws.freeze_panes = "A2"
    # Adiciona os filtros automáticos no cabeçalho
    ws.auto_filter.ref = ws.dimensions

def ajustar_largura(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = adjusted_width if adjusted_width > 15 else 15

# --- MOTOR PRINCIPAL DO EXCEL ---

def gerar_planilha(dados):
    wb = openpyxl.Workbook()
    FORMATO_MOEDA = 'R$ #,##0.00'
    
    # ---------------------------------------------------------
    # ABA 1: RESUMO EXECUTIVO (Opcional, mas dá um toque premium)
    # ---------------------------------------------------------
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Executivo"
    ws_resumo.sheet_view.showGridLines = False # Esconde as linhas de grade para parecer um Dashboard
    
    ws_resumo["B2"] = "DIAGNÓSTICO FISCAL CONSOLIDADO"
    ws_resumo["B2"].font = Font(size=16, bold=True, color="0F172A")
    
    ws_resumo["B4"] = "Razão Social:"
    ws_resumo["C4"] = dados['empresa'].get('razao_social', 'N/A')
    ws_resumo["C4"].font = Font(bold=True)
    
    ws_resumo["B5"] = "CNPJ:"
    ws_resumo["C5"] = dados['empresa'].get('cnpj', 'N/A')
    
    ws_resumo["B6"] = "Situação:"
    ws_resumo["C6"] = dados['empresa'].get('situacao', 'N/A')
    
    ws_resumo.column_dimensions["B"].width = 20
    ws_resumo.column_dimensions["C"].width = 50

    # ---------------------------------------------------------
    # ABA 2: PENDÊNCIAS RFB (A mais detalhada de todas)
    # ---------------------------------------------------------
    ws_pend = wb.create_sheet(title="Pendências RFB")
    cabecalhos_pend = ["Órgão", "Natureza / Tributo", "Período Apuração", "Valor Original", "Multa", "Juros", "Saldo Consolidado", "Status"]
    estilizar_cabecalho(ws_pend, cabecalhos_pend)
    
    row_idx = 2
    for p in dados.get('pendencias', []):
        # Converte as strings extraídas para números reais antes de jogar no Excel
        v_orig = converter_para_numero(p['vl_original'])
        v_multa = converter_para_numero(p['multa'])
        v_juros = converter_para_numero(p['juros'])
        v_cons = converter_para_numero(p['vl_consolidado'])
        
        ws_pend.append([p['orgao'], p['tipo'], p['periodo'], v_orig, v_multa, v_juros, v_cons, p['status']])
        
        # Formatação das colunas monetárias (D, E, F, G)
        for col_letter in ['D', 'E', 'F', 'G']:
            ws_pend[f"{col_letter}{row_idx}"].number_format = FORMATO_MOEDA
        
        # Formatação condicional visual para o Status
        status_cell = ws_pend[f"H{row_idx}"]
        status_cell.font = Font(bold=True)
        if "DEVEDOR" in p['status'].upper():
            status_cell.font = Font(color="991B1B", bold=True) # Vermelho
            status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        elif "ATRASO" in p['status'].upper():
            status_cell.font = Font(color="9A3412", bold=True) # Laranja
            status_cell.fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
            
        row_idx += 1

    # Adiciona a linha de TOTAL automática se houver débitos
    if row_idx > 2:
        ws_pend[f"C{row_idx}"] = "TOTAL GERAL:"
        ws_pend[f"C{row_idx}"].font = Font(bold=True)
        ws_pend[f"C{row_idx}"].alignment = Alignment(horizontal="right")
        
        # Insere Fórmulas de SOMA do Excel nas colunas financeiras
        for col_letter in ['D', 'E', 'F', 'G']:
            ws_pend[f"{col_letter}{row_idx}"] = f"=SUM({col_letter}2:{col_letter}{row_idx-1})"
            ws_pend[f"{col_letter}{row_idx}"].number_format = FORMATO_MOEDA
            ws_pend[f"{col_letter}{row_idx}"].font = Font(bold=True)
            ws_pend[f"{col_letter}{row_idx}"].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    ajustar_largura(ws_pend)

    # ---------------------------------------------------------
    # ABA 3: DÍVIDA ATIVA PGFN
    # ---------------------------------------------------------
    ws_pgfn = wb.create_sheet(title="Dívida Ativa PGFN")
    estilizar_cabecalho(ws_pgfn, ["Inscrição PGFN", "Origem / Tributo", "Situação"])
    
    if not dados.get('pgfn'):
        ws_pgfn.append(["-", "Nenhuma pendência PGFN detectada", "-"])
    else:
        for pgfn in dados.get('pgfn', []):
            ws_pgfn.append([pgfn['inscricao'], pgfn['tributo'], pgfn['status']])
            # Destaca inscrições em dívida ativa em vermelho
            ws_pgfn[f"C{ws_pgfn.max_row}"].font = Font(color="991B1B", bold=True)
            
    ajustar_largura(ws_pgfn)

    # ---------------------------------------------------------
    # ABA 4: HISTÓRICO SIMPLES NACIONAL
    # ---------------------------------------------------------
    ws_simples = wb.create_sheet(title="Simples Nacional")
    estilizar_cabecalho(ws_simples, ["Evento", "Data Inclusão", "Data Exclusão"])
    
    if not dados.get('simples_nacional'):
        ws_simples.append(["-", "Sem histórico detectado", "-"])
    else:
        for s in dados.get('simples_nacional', []):
            ws_simples.append([s['evento'], s['inclusao'], s['exclusao']])
    ajustar_largura(ws_simples)

    # ---------------------------------------------------------
    # ABA 5: CERTIDÕES
    # ---------------------------------------------------------
    ws_cert = wb.create_sheet(title="Certidões Emitidas")
    estilizar_cabecalho(ws_cert, ["Tipo de Certidão", "Código de Controle", "Data de Emissão", "Validade"])
    
    if not dados.get('certidoes'):
        ws_cert.append(["-", "Nenhuma certidão localizada", "-", "-"])
    else:
        for c in dados.get('certidoes', []):
            ws_cert.append([c['tipo'], c['codigo'], c['emissao'], c['validade']])
    ajustar_largura(ws_cert)

    # ---------------------------------------------------------
    # ABA 6: QUADRO SOCIETÁRIO
    # ---------------------------------------------------------
    ws_socios = wb.create_sheet(title="Quadro Societário")
    estilizar_cabecalho(ws_socios, ["CPF/CNPJ", "Nome do Sócio", "Qualificação"])
    
    if not dados.get('socios'):
        ws_socios.append(["-", "Nenhum sócio identificado", "-"])
    else:
        for s in dados.get('socios', []):
            ws_socios.append([s['documento'], s['nome'], s['qualificacao']])
    ajustar_largura(ws_socios)

    # SALVAMENTO
    # Limpa caracteres especiais do CNPJ para o nome do arquivo
    cnpj_limpo = dados['empresa'].get('cnpj', '000').replace('/', '').replace('.', '').replace('-', '')
    nome_arquivo = f"SitFiscal_{cnpj_limpo}.xlsx"
    wb.save(nome_arquivo)
    
    return nome_arquivo